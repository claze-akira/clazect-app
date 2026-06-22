import streamlit as st
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
import openpyxl

# ===== 設定 =====
SPREADSHEET_ID = '1h06FfSGadEqViz77rReSlbIs_QIryOE1JloWqjR2GCU'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']

EXPENSE_KW = ['旅費交通費','支払手数料','交際費','通信費','消耗品費','リース料','研修費','会議費',
              '地代家賃','外注費','広告宣伝費','水道光熱費','修繕費','保険料','新聞図書費',
              '福利厚生費','雑費','給与手当','給料手当','役員報酬','法定福利費','減価償却費',
              '租税公課','システム利用料','採用費','業務委託費','接待交際費','支払報酬料',
              '支払報酬','諸会費','寄付金','車両費','賞与','退職給与']
SALES_KW = ['売上高']
COGS_KW = ['仕入高', '売上原価']
MONTH_MAP = {'9月':'09','10月':'10','11月':'11','12月':'12','1月':'01','2月':'02',
             '3月':'03','4月':'04','5月':'05','6月':'06','7月':'07','8月':'08'}

# ===== Google Sheets接続 =====
@st.cache_resource
def get_gsheet_client():
    try:
        creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=SCOPES)
    except:
        creds = Credentials.from_service_account_file('credentials.json', scopes=SCOPES)
    return gspread.authorize(creds)

def get_spreadsheet():
    return get_gsheet_client().open_by_key(SPREADSHEET_ID)

def save_to_sheet(sh, sheet_name, df):
    try:
        ws = sh.worksheet(sheet_name)
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=sheet_name, rows=5000, cols=50)
    ws.update([df.columns.tolist()] + df.fillna('').values.tolist())

def load_from_sheet(sh, sheet_name):
    try:
        ws = sh.worksheet(sheet_name)
        data = ws.get_all_records()
        return pd.DataFrame(data) if data else None
    except:
        return None

# ===== 仕訳帳CSVパース =====
def parse_jn_csv(uploaded_file):
    try:
        content = uploaded_file.read()
        for enc in ('cp932', 'shift-jis', 'utf-8-sig', 'utf-8'):
            try:
                text = content.decode(enc)
                break
            except:
                continue
        else:
            return None, '文字コードを判定できませんでした'

        import csv, io
        lines = [l for l in text.split('\n') if l.strip()]
        if len(lines) < 2:
            return None, 'データが少なすぎます'

        reader = csv.reader(io.StringIO(lines[0]))
        header = next(reader)

        def gi(name):
            try: return header.index(name)
            except: return -1

        iDate=gi('取引日'); iDkAcc=gi('借方勘定科目'); iDkAmt=gi('借方金額')
        iDkDept=gi('借方部門'); iCrAcc=gi('貸方勘定科目'); iCrAmt=gi('貸方金額')
        iCrDept=gi('貸方部門'); iDkPartner=gi('借方取引先名'); iNote=gi('取引内容')

        if iDkAcc < 0 or iDkAmt < 0:
            return None, '仕訳帳CSVの列が見つかりません'

        rows = []
        for line in lines[1:]:
            r = next(csv.reader(io.StringIO(line)))
            if not r: continue
            date_raw = r[iDate].strip() if iDate >= 0 and iDate < len(r) else ''
            if not date_raw or date_raw == 'NaN': continue
            month = date_raw[:7]

            dk_acc = r[iDkAcc].strip() if iDkAcc < len(r) else ''
            try:
                dk_amt = float(r[iDkAmt].replace(',','').replace('¥','')) if iDkAmt < len(r) and r[iDkAmt].strip() else 0
            except: dk_amt = 0
            dk_dept = r[iDkDept].strip() if iDkDept >= 0 and iDkDept < len(r) else ''
            dk_dept = '' if dk_dept in ('NaN','') else dk_dept
            cr_acc = r[iCrAcc].strip() if iCrAcc >= 0 and iCrAcc < len(r) else ''
            try:
                cr_amt = float(r[iCrAmt].replace(',','').replace('¥','')) if iCrAmt >= 0 and iCrAmt < len(r) and r[iCrAmt].strip() else 0
            except: cr_amt = 0
            cr_dept = r[iCrDept].strip() if iCrDept >= 0 and iCrDept < len(r) else ''
            cr_dept = '' if cr_dept in ('NaN','') else cr_dept
            partner = r[iDkPartner].strip() if iDkPartner >= 0 and iDkPartner < len(r) else ''
            partner = '' if partner == 'NaN' else partner
            note = r[iNote].strip()[:40] if iNote >= 0 and iNote < len(r) else ''
            note = '' if note == 'NaN' else note

            if any(k in cr_acc for k in SALES_KW) and cr_amt > 0:
                rows.append({'date':date_raw,'month':month,'type':'売上高','account':cr_acc,
                             'dept':cr_dept,'amount':cr_amt,'partner':partner,'note':note})
            if any(k in dk_acc for k in COGS_KW) and dk_amt > 0:
                rows.append({'date':date_raw,'month':month,'type':'仕入高','account':dk_acc,
                             'dept':dk_dept,'amount':dk_amt,'partner':partner,'note':note})
            if any(k in dk_acc for k in EXPENSE_KW) and dk_amt > 0:
                rows.append({'date':date_raw,'month':month,'type':'費用','account':dk_acc,
                             'dept':dk_dept,'amount':dk_amt,'partner':partner,'note':note})

        return pd.DataFrame(rows), None
    except Exception as e:
        return None, str(e)

# ===== 予算xlsxパース =====
def parse_budget_xlsx(uploaded_file):
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        bud_rows = []
        base_year = 2025

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            dept = sheet_name

            # ヘッダー行を探す（月ラベルがある行）
            header_row = None
            month_cols = {}
            for row in ws.iter_rows(min_row=1, max_row=10):
                for cell in row:
                    val = str(cell.value or '').strip()
                    if val in MONTH_MAP:
                        header_row = cell.row
                        mm = MONTH_MAP[val]
                        yyyy = base_year if int(mm) >= 9 else base_year + 1
                        month_cols[cell.column] = f'{yyyy}-{mm}'
            if not header_row or not month_cols:
                continue

            # データ行を走査
            for row in ws.iter_rows(min_row=header_row+1):
                label = str(row[0].value or '').strip()
                if not label or label.startswith('▼') or label.startswith('#'):
                    continue
                if label in ('費用合計', '売上総利益', '営業利益', '年間合計'):
                    continue
                for col, month in month_cols.items():
                    cell = ws.cell(row=row[0].row, column=col)
                    try:
                        val = float(cell.value or 0)
                    except:
                        val = 0
                    if val == 0:
                        continue
                    if label == '売上高':
                        bud_type = '売上高'
                    elif label == '仕入高':
                        bud_type = '仕入高'
                    else:
                        bud_type = '費用'
                    bud_rows.append({'dept':dept,'month':month,'type':bud_type,'account':label,'amount':val})

        return pd.DataFrame(bud_rows) if bud_rows else None, None
    except Exception as e:
        return None, str(e)

# ===== 集計 =====
def aggregate(df, dept=None, month=None):
    d = df.copy()
    if dept and dept != '全体': d = d[d['dept']==dept]
    if month and month != '累計': d = d[d['month']==month]
    sales = d[d['type']=='売上高']['amount'].sum()
    cogs  = d[d['type']=='仕入高']['amount'].sum()
    exp   = d[d['type']=='費用']['amount'].sum()
    return {'sales':sales,'cogs':cogs,'expense':exp,'gross':sales-cogs,'op':sales-cogs-exp}

def agg_bud(bdf, dept=None, month=None):
    if bdf is None: return {'sales':0,'cogs':0,'expense':0,'gross':0,'op':0}
    d = bdf.copy()
    if dept and dept != '全体': d = d[d['dept']==dept]
    if month and month != '累計': d = d[d['month']==month]
    sales = d[d['type']=='売上高']['amount'].sum()
    cogs  = d[d['type']=='仕入高']['amount'].sum()
    exp   = d[d['type']=='費用']['amount'].sum()
    return {'sales':sales,'cogs':cogs,'expense':exp,'gross':sales-cogs,'op':sales-cogs-exp}

def fmt(n):
    if n == 0: return '¥0'
    sign = '-' if n < 0 else ''
    return f"{sign}¥{abs(int(n)):,}"

def pct(a, b):
    return f"{a/b*100:.1f}%" if b else '-'

def diff_str(act, bud):
    if bud == 0: return ''
    d = act - bud
    arrow = '▲' if d >= 0 else '▼'
    return f"{arrow} {fmt(abs(d))}"

# ===== UI =====
st.set_page_config(page_title='業績分析表', layout='wide', page_icon='📈')
st.title('📈 業績分析表')

# ===== サイドバー =====
with st.sidebar:
    st.header('📂 データ管理')

    st.subheader('実績（仕訳帳CSV）')
    jn_file = st.file_uploader('freee仕訳帳CSV（新形式）', type='csv', key='jn')
    if jn_file:
        with st.spinner('取込中...'):
            df_new, err = parse_jn_csv(jn_file)
            if err:
                st.error(err)
            else:
                try:
                    sh = get_spreadsheet()
                    save_to_sheet(sh, '実績データ', df_new)
                    st.success(f'✅ {len(df_new)}件を保存しました')
                    st.cache_data.clear()
                except Exception as e:
                    st.error(f'保存エラー: {e}')

    st.divider()
    st.subheader('予算（xlsx）')
    bud_file = st.file_uploader('予算入力テンプレート xlsx', type=['xlsx','xls'], key='bud')
    if bud_file:
        with st.spinner('取込中...'):
            bdf_new, err = parse_budget_xlsx(bud_file)
            if err:
                st.error(err)
            elif bdf_new is None:
                st.warning('予算データが見つかりませんでした')
            else:
                try:
                    sh = get_spreadsheet()
                    save_to_sheet(sh, '予算データ', bdf_new)
                    st.success(f'✅ 予算データを保存しました（{len(bdf_new)}件）')
                    st.cache_data.clear()
                except Exception as e:
                    st.error(f'保存エラー: {e}')

    st.divider()
    st.caption('データはGoogle Sheetsに自動保存されます')

# ===== データ読込 =====
@st.cache_data(ttl=60)
def load_data():
    try:
        sh = get_spreadsheet()
        df = load_from_sheet(sh, '実績データ')
        if df is not None and not df.empty:
            df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0)
            return df
    except: pass
    return None

@st.cache_data(ttl=60)
def load_budget():
    try:
        sh = get_spreadsheet()
        bdf = load_from_sheet(sh, '予算データ')
        if bdf is not None and not bdf.empty:
            bdf['amount'] = pd.to_numeric(bdf['amount'], errors='coerce').fillna(0)
            return bdf
    except: pass
    return None

df = load_data()
bdf = load_budget()
has_bud = bdf is not None and not bdf.empty

if df is None or df.empty:
    st.info('👈 サイドバーから仕訳帳CSVを取り込んでください')
    st.stop()

# ===== フィルター =====
months = sorted(df['month'].unique().tolist())
depts  = sorted(df[df['dept'] != '']['dept'].unique().tolist())

col1, col2 = st.columns([3, 1])
with col1:
    month_options = months + ['累計']
    sel_month = st.selectbox('月', month_options, index=len(month_options)-2)
with col2:
    dept_options = ['全体'] + depts
    sel_dept = st.selectbox('部門', dept_options)

st.divider()

# ===== サマリーカード =====
agg  = aggregate(df, sel_dept, sel_month)
bagg = agg_bud(bdf, sel_dept, sel_month)

def metric_with_budget(label, act, bud, help_text=''):
    delta = None
    if has_bud and bud != 0:
        d = act - bud
        delta = f"予算比 {'+' if d>=0 else ''}{fmt(d)}"
    st.metric(label, fmt(act), delta, help=help_text if help_text else None)

c1, c2, c3, c4 = st.columns(4)
with c1: metric_with_budget('売上高', agg['sales'], bagg['sales'])
with c2: metric_with_budget('売上総利益', agg['gross'], bagg['gross'], f"粗利率 {pct(agg['gross'],agg['sales'])}")
with c3: metric_with_budget('費用合計', agg['expense'], bagg['expense'])
with c4: metric_with_budget('営業利益', agg['op'], bagg['op'], f"利益率 {pct(agg['op'],agg['sales'])}")

st.divider()

# ===== タブ =====
tab1, tab2, tab3 = st.tabs(['📊 費用内訳', '📅 月次詳細', '📈 推移グラフ'])

with tab1:
    d = df.copy()
    if sel_dept != '全体': d = d[d['dept']==sel_dept]
    if sel_month != '累計': d = d[d['month']==sel_month]

    exp_df = d[d['type']=='費用'].groupby('account')['amount'].sum().reset_index()
    exp_df.columns = ['勘定科目', '実績']
    exp_df = exp_df[exp_df['実績']>0].sort_values('実績', ascending=False)

    if has_bud:
        bd = bdf.copy()
        if sel_dept != '全体': bd = bd[bd['dept']==sel_dept]
        if sel_month != '累計': bd = bd[bd['month']==sel_month]
        bexp = bd[bd['type']=='費用'].groupby('account')['amount'].sum().reset_index()
        bexp.columns = ['勘定科目','予算']
        exp_df = exp_df.merge(bexp, on='勘定科目', how='left').fillna(0)
        exp_df['差額'] = exp_df.apply(lambda r: diff_str(r['実績'], r['予算']), axis=1)
        exp_df['構成比'] = exp_df['実績'].apply(lambda x: pct(x, agg['expense']))
        exp_df['予算'] = exp_df['予算'].apply(fmt)
    else:
        exp_df['構成比'] = exp_df['実績'].apply(lambda x: pct(x, agg['expense']))

    exp_df['実績'] = exp_df['実績'].apply(fmt)

    # 行選択で明細表示
    st.write('##### 科目をクリックすると明細を表示します')
    sel_row = st.dataframe(
        exp_df, use_container_width=True, hide_index=True,
        on_select='rerun', selection_mode='single-row'
    )

    # 選択された行の明細を表示
    if sel_row and sel_row.selection and sel_row.selection.rows:
        selected_acc = exp_df.iloc[sel_row.selection.rows[0]]['勘定科目']
        st.subheader(f'🔍 明細：{selected_acc}')
        detail = d[d['account']==selected_acc][['date','dept','partner','note','amount']].copy()
        detail = detail.sort_values('date')
        detail.columns = ['日付','部門','取引先','摘要','金額']
        detail['金額'] = detail['金額'].apply(fmt)
        st.dataframe(detail, use_container_width=True, hide_index=True)
        st.caption(f'{len(detail)}件')

with tab2:
    d = df.copy()
    if sel_dept != '全体': d = d[d['dept']==sel_dept]
    bd = bdf.copy() if has_bud else None
    if has_bud and sel_dept != '全体': bd = bd[bd['dept']==sel_dept]

    # 月フィルター（複数選択）
    month_labels_all = [m[5:]+'月' for m in months]
    sel_months_filter = st.multiselect(
        '月を絞り込む（複数選択可）', month_labels_all, default=month_labels_all, key='tab2_months'
    )
    # 選択された月のみ使用
    filtered_months = [m for m, ml in zip(months, month_labels_all) if ml in sel_months_filter]
    month_labels = [m[5:]+'月' for m in filtered_months]
    all_cols = ['項目'] + month_labels + ['累計']

    def get_act(type_, acc=None, m=None):
        dd = d[d['month']==m] if m else d
        dd = dd[dd['type']==type_]
        if acc: dd = dd[dd['account']==acc]
        return dd['amount'].sum()

    def get_bud(type_, acc=None, m=None):
        if not has_bud: return 0
        bb = bd[bd['month']==m] if m else bd
        bb = bb[bb['type']==type_]
        if acc: bb = bb[bb['account']==acc]
        return bb['amount'].sum()

    def make_row(label, vals, is_section=False, is_total=False, indent=False):
        prefix = '　' if indent else ''
        return {'項目': prefix + label, '_vals': vals, '_section': is_section, '_total': is_total}

    # 費用科目一覧（実績＋予算）
    exp_accs = sorted(set(
        d[d['type']=='費用']['account'].unique().tolist() +
        (bd[bd['type']=='費用']['account'].unique().tolist() if has_bud else [])
    ))

    # 損益計算書の行を構築
    def build_pl_rows():
        rows = []

        def add(label, act_fn, bud_fn, indent=False, is_total=False, is_section=False):
            act_m = [act_fn(m) for m in filtered_months]
            act_t = sum(act_m)
            bud_m = [bud_fn(m) for m in filtered_months]
            bud_t = sum(bud_m)
            vals_act = [fmt(v) for v in act_m] + [fmt(act_t)]
            vals_bud = [fmt(v) for v in bud_m] + [fmt(bud_t)]
            vals_diff = [diff_str(a, b) for a, b in zip(act_m, bud_m)] + [diff_str(act_t, bud_t)]
            rows.append({'項目': ('　' if indent else '') + label,
                         'vals_act': vals_act, 'vals_bud': vals_bud, 'vals_diff': vals_diff,
                         'is_section': is_section, 'is_total': is_total})

        add('売上高', lambda m: get_act('売上高', m=m), lambda m: get_bud('売上高', m=m))
        add('売上原価（仕入高）', lambda m: get_act('仕入高', m=m), lambda m: get_bud('仕入高', m=m))
        add('売上総利益', lambda m: get_act('売上高',m=m)-get_act('仕入高',m=m),
            lambda m: get_bud('売上高',m=m)-get_bud('仕入高',m=m), is_total=True)
        add('販売管理費', lambda m: 0, lambda m: 0, is_section=True)
        for acc in exp_accs:
            add(acc, lambda m, a=acc: get_act('費用', acc=a, m=m),
                lambda m, a=acc: get_bud('費用', acc=a, m=m), indent=True)
        add('費用合計', lambda m: get_act('費用',m=m), lambda m: get_bud('費用',m=m), is_total=True)
        add('営業利益', lambda m: get_act('売上高',m=m)-get_act('仕入高',m=m)-get_act('費用',m=m),
            lambda m: get_bud('売上高',m=m)-get_bud('仕入高',m=m)-get_bud('費用',m=m), is_total=True)
        return rows

    pl_rows = build_pl_rows()


    # ダイアログ定義（明細ポップアップ）
    @st.dialog('🔍 仕訳明細', width='large')
    def show_detail_dialog(acc, target_month, dept_data):
        title = f'{acc}　{target_month if target_month != "累計" else "累計（全期間）"}'
        st.write(f'**{title}**')
        dd = dept_data[dept_data['account']==acc].copy()
        if target_month != '累計':
            # 月ラベルから月コードに変換
            target_m = next((m for m, ml in zip(months, month_labels_all) if ml == target_month), None)
            if target_m:
                dd = dd[dd['month']==target_m]
        dd = dd[['date','month','dept','partner','note','amount']].sort_values('date')
        dd.columns = ['日付','月','部門','取引先','摘要','金額']
        total = dd['金額'].sum() if len(dd) > 0 else 0
        dd['金額'] = dd['金額'].apply(fmt)
        st.caption(f'{len(dd)}件　合計：{fmt(total)}')
        st.dataframe(dd, use_container_width=True, hide_index=True)

    # セッション状態で選択した科目・月を管理
    if 'dialog_acc' not in st.session_state:
        st.session_state.dialog_acc = None
        st.session_state.dialog_month = None

    # 損益計算書テーブルをボタン形式で表示
    st.write('##### 金額をクリックすると明細ウィンドウが開きます')
    header_cols = st.columns([3] + [1]*len(month_labels) + [1.2])
    header_cols[0].write('**項目**')
    for i, ml in enumerate(month_labels):
        header_cols[i+1].write(f'**{ml}**')
    header_cols[-1].write('**累計**')

    for row_idx, r in enumerate(pl_rows):
        if r['is_section']:
            st.markdown(f"**━━ {r['項目']} ━━**")
            continue
        cols = st.columns([3] + [1]*len(month_labels) + [1.2])
        label = r['項目']
        is_clickable = not r['is_total']
        cols[0].write(label)
        for i, (ml, act_v) in enumerate(zip(month_labels, r['vals_act'])):
            if is_clickable and act_v != '¥0':
                if cols[i+1].button(act_v, key=f'btn_{row_idx}_{i}', use_container_width=True):
                    st.session_state.dialog_acc = label.strip()
                    st.session_state.dialog_month = ml
            else:
                cols[i+1].write(act_v)
        # 累計
        cum_v = r['vals_act'][-1]
        if is_clickable and cum_v != '¥0':
            if cols[-1].button(cum_v, key=f'btn_{row_idx}_cum', use_container_width=True):
                st.session_state.dialog_acc = label.strip()
                st.session_state.dialog_month = '累計'
        else:
            cols[-1].write(cum_v)

    # ダイアログ表示
    if st.session_state.dialog_acc:
        show_detail_dialog(st.session_state.dialog_acc, st.session_state.dialog_month, d)
        st.session_state.dialog_acc = None
        st.session_state.dialog_month = None

    # 予実比較テーブル
    if has_bud:
        st.write('**予実比較（実績 vs 予算）**')
        comp_rows = []
        for r in pl_rows:
            if r['is_section']:
                comp_rows.append({col: ('━━ '+r['項目']+' ━━' if col=='項目' else '') for col in all_cols})
                continue
            row_d = {'項目': r['項目']}
            for i, ml in enumerate(month_labels):
                act_v = r['vals_act'][i]
                bud_v = r['vals_bud'][i]
                dif_v = r['vals_diff'][i]
                row_d[ml] = f"{act_v} / {bud_v} {dif_v}"
            row_d['累計'] = f"{r['vals_act'][-1]} / {r['vals_bud'][-1]} {r['vals_diff'][-1]}"
            comp_rows.append(row_d)
        comp_df = pd.DataFrame(comp_rows)
        st.dataframe(comp_df, use_container_width=True, hide_index=True)

with tab3:
    d = df.copy()
    if sel_dept != '全体': d = d[d['dept']==sel_dept]

    chart_data = []
    for m in months:
        md = d[d['month']==m]
        s = md[md['type']=='売上高']['amount'].sum()
        c = md[md['type']=='仕入高']['amount'].sum()
        e = md[md['type']=='費用']['amount'].sum()
        row = {'月': m[5:]+'月', '売上高（実績）': s, '売上総利益': s-c, '営業利益（実績）': s-c-e}
        if has_bud:
            bd = bdf.copy()
            if sel_dept != '全体': bd = bd[bd['dept']==sel_dept]
            bm = bd[bd['month']==m]
            bs = bm[bm['type']=='売上高']['amount'].sum()
            bc = bm[bm['type']=='仕入高']['amount'].sum()
            be = bm[bm['type']=='費用']['amount'].sum()
            row['売上高（予算）'] = bs
            row['営業利益（予算）'] = bs-bc-be
        chart_data.append(row)

    chart_df = pd.DataFrame(chart_data).set_index('月')
    cols_to_show = ['売上高（実績）','売上総利益','営業利益（実績）']
    if has_bud: cols_to_show += ['売上高（予算）','営業利益（予算）']
    st.bar_chart(chart_df[cols_to_show])
